import os
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# -----------------------------
# Config
# -----------------------------
ticker_file = "nifty50_list.xlsx"
download_dir = "downloads"
clean_dir = "clean_data"
backtest_dir = "backtest"
summary_file = "Master_Summary.xlsx"

start_date = "2025-01-01"
end_date = "2025-12-31"
interval = "1h"

initial_capital = 100000  # Starting capital

# Ensure folders exist
for folder in [download_dir, clean_dir, backtest_dir]:
    os.makedirs(folder, exist_ok=True)

# -----------------------------
# Load Tickers
# -----------------------------
tickers_df = pd.read_excel(ticker_file)
tickers = tickers_df.iloc[:, 0].dropna().tolist()

master_results = []

# -----------------------------
# Process Each Ticker
# -----------------------------
for symbol in tickers:
    try:
        print(f"\n📥 Fetching data for {symbol}...")

        # --- Download Data ---
        data = yf.download(
            symbol,
            start=start_date,
            end=end_date,
            interval=interval,
            auto_adjust=True
        )

        if data.empty:
            print(f"⚠️ No data for {symbol}, skipping...")
            continue

        # Flatten columns if multi-index
        if isinstance(data.columns, pd.MultiIndex):
            data.columns = [col[0] if isinstance(col, tuple) else col for col in data.columns]

        # Convert to IST
        if data.index.tz is not None:
            data.index = data.index.tz_convert("Asia/Kolkata")
        data.index = data.index.tz_localize(None)

        # --- Clean Data ---
        data = data.reset_index()
        data["Date"] = data["Datetime"].dt.date
        data["Time"] = data["Datetime"].dt.time
        data = data[["Date", "Time", "Close", "High", "Low", "Open", "Volume"]]

        clean_file = os.path.join(clean_dir, f"{symbol}_1H_Data_Clean.xlsx")
        if os.path.exists(clean_file):
            os.remove(clean_file)
        data.to_excel(clean_file, index=False)
        print(f"✅ Cleaned file saved → {clean_file}")

        # -----------------------------
        # Backtest with Compounded Capital & Quantity
        # -----------------------------
        first_candles = data.groupby("Date").first().reset_index()
        results = []
        current_capital = initial_capital

        for _, row in first_candles.iterrows():
            day = row["Date"]
            day_data = data[data["Date"] == day]

            first_high = row["High"]
            first_low = row["Low"]
            last_close = day_data.iloc[-1]["Close"]

            trend = "Neutral"
            entry = None
            exit_price = last_close
            pnl = 0
            invested_qty = 0

            # Determine daily trade and quantity
            if (day_data["High"] > first_high).any():
                entry = first_high
                invested_qty = int(current_capital / entry)
                trend = "Upside Follow" if last_close > first_high else "Upside Fake"
                pnl = (last_close - entry) * invested_qty

            elif (day_data["Low"] < first_low).any():
                entry = first_low
                invested_qty = int(current_capital / entry)
                trend = "Downside Follow" if last_close < first_low else "Downside Fake"
                pnl = (entry - last_close) * invested_qty

            current_capital += pnl  # update capital after trade

            results.append({
                "Date": day,
                "Trend": trend,
                "Entry": entry,
                "Exit": exit_price,
                "Quantity": invested_qty,
                "PnL": pnl,
                "Capital": current_capital
            })

        trend_df = pd.DataFrame(results)
        final_capital = current_capital
        total_profit = final_capital - initial_capital

        summary = trend_df["Trend"].value_counts().to_dict()
        total_trades = len(trend_df)
        follow_trades = summary.get("Upside Follow", 0) + summary.get("Downside Follow", 0)
        win_rate = (follow_trades / total_trades * 100) if total_trades > 0 else 0

        # -----------------------------
        # Save Backtest file with Highlights
        # -----------------------------
        backtest_file = os.path.join(backtest_dir, f"{symbol}_Backtest.xlsx")
        trend_df.to_excel(backtest_file, index=False, sheet_name="Backtest")

        wb = load_workbook(backtest_file)
        ws = wb["Backtest"]

        green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        # Find column indices
        pnl_col = None
        capital_col = None
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header == "PnL":
                pnl_col = col
            elif header == "Capital":
                capital_col = col

        # Apply highlights
        for row_num in range(2, ws.max_row + 1):
            # PnL highlight
            if pnl_col:
                pnl_value = ws.cell(row=row_num, column=pnl_col).value
                if pnl_value is not None:
                    ws.cell(row=row_num, column=pnl_col).fill = green_fill if pnl_value > 0 else red_fill
            # Capital highlight
            if capital_col:
                capital_value = ws.cell(row=row_num, column=capital_col).value
                prev_capital = initial_capital if row_num == 2 else ws.cell(row=row_num - 1, column=capital_col).value
                if capital_value is not None:
                    ws.cell(row=row_num, column=capital_col).fill = green_fill if capital_value > prev_capital else red_fill

        wb.save(backtest_file)
        print(f"📊 Backtest saved with PnL & Capital highlights → {backtest_file}")

        # -----------------------------
        # Master Results
        # -----------------------------
        master_results.append({
            "Ticker": symbol,
            "Total Trades": total_trades,
            "Upside Follow": summary.get("Upside Follow", 0),
            "Downside Follow": summary.get("Downside Follow", 0),
            "Upside Fake": summary.get("Upside Fake", 0),
            "Downside Fake": summary.get("Downside Fake", 0),
            "Win Rate %": round(win_rate, 2),
            "Final Capital": round(final_capital, 2),
            "Profit": round(total_profit, 2),
            "Total Profit (Year End)": round(final_capital - initial_capital, 2)  # <-- Added
        })

    except Exception as e:
        print(f"❌ Error for {symbol}: {e}")

# -----------------------------
# Save Master Summary + Highlights
# -----------------------------
if master_results:
    master_df = pd.DataFrame(master_results)

    best_stock = master_df.loc[master_df["Final Capital"].idxmax()]
    worst_stock = master_df.loc[master_df["Final Capital"].idxmin()]

    with pd.ExcelWriter(summary_file, engine="openpyxl") as writer:
        master_df.to_excel(writer, index=False, sheet_name="Summary")
        ws = writer.book["Summary"]

        green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        # Highlight Win Rate %
        win_rate_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "Win Rate %":
                win_rate_col = col
                break

        if win_rate_col:
            for row in range(2, ws.max_row + 1):
                value = ws.cell(row=row, column=win_rate_col).value
                if value is not None:
                    if value > 60:
                        ws.cell(row=row, column=win_rate_col).fill = green_fill
                    elif value < 40:
                        ws.cell(row=row, column=win_rate_col).fill = red_fill

        # Append Best & Worst performers
        ws.append([])
        ws.append([
            "Best Performer",
            best_stock["Ticker"],
            f"Capital: {best_stock['Final Capital']}",
            f"Profit: {best_stock['Profit']}",
            f"Win Rate: {best_stock['Win Rate %']}%",
            f"Total Profit (Year End): {best_stock['Total Profit (Year End)']}"  # <-- Added
        ])
        ws.append([
            "Worst Performer",
            worst_stock["Ticker"],
            f"Capital: {worst_stock['Final Capital']}",
            f"Profit: {worst_stock['Profit']}",
            f"Win Rate: {worst_stock['Win Rate %']}%",
            f"Total Profit (Year End): {worst_stock['Total Profit (Year End)']}"  # <-- Added
        ])

    print(f"\n✅ Master Summary saved → {summary_file}")

    # Console Summary
    print("\n================= 📊 SUMMARY REPORT =================")
    print(master_df[["Ticker", "Total Trades", "Win Rate %", "Final Capital", "Profit", "Total Profit (Year End)"]].to_string(index=False))
    print("------------------------------------------------------")
    print(f"🏆 Best Performer : {best_stock['Ticker']} | Capital: ₹{best_stock['Final Capital']} | Profit: ₹{best_stock['Profit']} | Total Profit: ₹{best_stock['Total Profit (Year End)']} | Win Rate: {best_stock['Win Rate %']}%")
    print(f"⚠️ Worst Performer: {worst_stock['Ticker']} | Capital: ₹{worst_stock['Final Capital']} | Profit: ₹{worst_stock['Profit']} | Total Profit: ₹{worst_stock['Total Profit (Year End)']} | Win Rate: {worst_stock['Win Rate %']}%")
    print("======================================================")
else:
    print("⚠️ No tickers processed successfully. Master Summary not created.")
